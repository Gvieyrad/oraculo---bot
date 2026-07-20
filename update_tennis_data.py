"""Parse tennis-data.co.uk XLSX files for 2025-2026 and update Elo."""
import json
import os
import sys
from datetime import datetime

import requests
from openpyxl import load_workbook

CACHE_DIR = "/home/noc/oraculo_v2/.oraculo_cache/tennis"
headers = {"User-Agent": "Mozilla/5.0"}
surface_map = {"Hard": "hard", "Clay": "clay", "Grass": "grass", "Carpet": "hard"}

for year in [2025, 2026]:
    url = "http://www.tennis-data.co.uk/{}/{}.xlsx".format(year, year)
    print("Downloading {}...".format(url))
    r = requests.get(url, headers=headers, timeout=30)
    if r.status_code != 200:
        print("  Failed: HTTP {}".format(r.status_code))
        continue

    xlsx_path = os.path.join(CACHE_DIR, "td_{}.xlsx".format(year))
    with open(xlsx_path, "wb") as f:
        f.write(r.content)

    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else "" for h in rows[0]]
    print("  Columns:", header[:15])

    matches = []
    for row in rows[1:]:
        d = dict(zip(header, row))
        winner = d.get("Winner", "")
        loser = d.get("Loser", "")
        surface = d.get("Surface", "Hard")
        date = d.get("Date", "")
        tourney = d.get("Tournament", d.get("ATP", ""))

        if not winner or not loser:
            continue

        if hasattr(date, "strftime"):
            date = date.strftime("%Y-%m-%d")
        else:
            date = str(date)[:10]

        matches.append({
            "winner": str(winner).strip(),
            "loser": str(loser).strip(),
            "date": date,
            "surface": surface_map.get(str(surface).strip(), "hard"),
            "tourney": str(tourney).strip() if tourney else "",
        })

    json_path = os.path.join(CACHE_DIR, "atp_{}.json".format(year))
    with open(json_path, "w") as f:
        json.dump(matches, f)

    dates = [m["date"] for m in matches if m["date"]]
    min_d = min(dates) if dates else "?"
    max_d = max(dates) if dates else "?"
    print("  {} matches ({} to {})".format(len(matches), min_d, max_d))

    recent = sorted(matches, key=lambda x: x["date"], reverse=True)[:5]
    print("  Most recent:")
    for m in recent:
        print("    {} {} beat {} ({})".format(m["date"], m["winner"], m["loser"], m["surface"]))
    wb.close()

# Test updated Elo
print("\n=== ELO WITH ALL DATA ===")
sys.path.insert(0, "/home/noc/oraculo_v2")
from oraculo_tennis import TennisElo

elo = TennisElo()
total = 0
for fname in sorted(os.listdir(CACHE_DIR)):
    if fname.endswith(".json"):
        data = json.load(open(os.path.join(CACHE_DIR, fname)))
        elo.process_matches(data)
        total += len(data)
        dates = [m.get("date", "") for m in data if m.get("date")]
        max_d = max(dates) if dates else "?"
        print("  {}: {} matches (to {})".format(fname, len(data), max_d))

print("\n  TOTAL: {} matches, {} players".format(total, len(elo.overall)))

print("\nTop 20:")
for name, rating in elo.get_top(20):
    print("  {:30s} {:.0f}".format(name, rating))

print("\nOur bet players (UPDATED):")
players = ["Arthur Fils", "Taylor Fritz", "Sebastian Korda", "Frances Tiafoe",
           "Ugo Humbert", "Quentin Halys", "Jiri Lehecka", "Martin Landaluce",
           "Francisco Cerundolo"]
for p in players:
    rating = elo.overall.get(p, 1500)
    matches_count = elo._match_count.get(p, 0)
    print("  {:25s} Elo={:.0f} ({} matches)".format(p, rating, matches_count))




# ── WTA Sackmann refresh ─────────────────────────────────────────────────────
def _refresh_wta(cache_dir=CACHE_DIR):
    """Download latest WTA results from tennis-data.co.uk, rebuild WTA Elo.

    2026-07-13: JeffSackmann/tennis_wta (and tennis_atp) disappeared from
    GitHub entirely (confirmed via GitHub API: account now has only 1 public
    repo, tennis_wta returns 404). Silently frozen since 2026-04-21 while
    real WTA bets kept getting placed on stale Elo. Switched to
    tennis-data.co.uk (same site ATP already uses successfully via td_*.xlsx
    in this file) -- {year}w/{year}.xlsx is the WTA equivalent, same column
    layout (Winner/Loser/Surface/Date/Tournament).
    """
    smap = {'Hard': 'hard', 'Clay': 'clay', 'Grass': 'grass', 'Carpet': 'hard'}
    updated = False
    for year in [2025, 2026]:
        url = 'http://www.tennis-data.co.uk/{0}w/{0}.xlsx'.format(year)
        try:
            r = requests.get(url, headers=headers, timeout=30)
            if r.status_code != 200:
                continue
            xlsx_path = os.path.join(cache_dir, 'td_wta_{}.xlsx'.format(year))
            with open(xlsx_path, 'wb') as f:
                f.write(r.content)
            wb = load_workbook(xlsx_path, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            header = [str(h).strip() if h else '' for h in rows[0]]
            matches = []
            for row in rows[1:]:
                d = dict(zip(header, row))
                winner = d.get('Winner', '')
                loser = d.get('Loser', '')
                if not winner or not loser:
                    continue
                date = d.get('Date', '')
                date = date.strftime('%Y-%m-%d') if hasattr(date, 'strftime') else str(date)[:10]
                matches.append({
                    'winner': str(winner).strip(), 'loser': str(loser).strip(),
                    'surface': smap.get(str(d.get('Surface', 'Hard')).strip(), 'hard'),
                    'date': date, 'tourney': str(d.get('Tournament', '') or '').strip(),
                    'tour': 'wta_real',
                })
            wb.close()
            jpath = os.path.join(cache_dir, 'wta_real_{}.json'.format(year))
            with open(jpath, 'w') as f:
                json.dump(matches, f)
            updated = True
        except Exception:
            pass
    if updated:
        pkl = os.path.join(cache_dir, 'elo_wta.pkl')
        if os.path.exists(pkl):
            os.remove(pkl)
    return updated


def update_cache():
    """Called by runner daily. Refreshes WTA data. Returns True on success."""
    try:
        return _refresh_wta()
    except Exception:
        return False
