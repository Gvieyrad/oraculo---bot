#!/usr/bin/env python3
"""
Tennis shadow pick resolver.
Uses tennis-data.co.uk XLSX files for set-by-set scores.
Resolves: Winner, Exact sets, Sets O/U, W+Total, wins-set markets.
"""
import os, sys, sqlite3, logging, re
from datetime import datetime, timedelta
from difflib import SequenceMatcher

SIBILA_DB  = '/home/noc/oraculo_v2/sibila.db'
CACHE_DIR  = '/home/noc/oraculo_v2/.oraculo_cache/tennis'
ATP_XLSX   = os.path.join(CACHE_DIR, 'td_2026.xlsx')
WTA_XLSX   = os.path.join(CACHE_DIR, 'wta_2026.xlsx')

log = logging.getLogger('tennis_resolver')
if not log.handlers:
    logging.basicConfig(level=logging.INFO, format='%(asctime)s %(message)s')

_XLSX_CACHE = {}


def _load_xlsx(path):
    if path in _XLSX_CACHE:
        return _XLSX_CACHE[path]
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(h).strip() if h else '' for h in rows[0]]
        matches = []
        for row in rows[1:]:
            d = dict(zip(header, row))
            winner = str(d.get('Winner') or '').strip()
            loser  = str(d.get('Loser')  or '').strip()
            if not winner or not loser or winner == 'None':
                continue
            date_val = d.get('Date')
            if hasattr(date_val, 'strftime'):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val or '')[:10]

            # Parse set scores
            sets = []
            for i in range(1, 6):
                wg = d.get(f'W{i}')
                lg = d.get(f'L{i}')
                try:
                    wg = int(wg) if wg not in (None, '', '\t') else None
                    lg = int(lg) if lg not in (None, '', '\t') else None
                except (ValueError, TypeError):
                    wg = lg = None
                if wg is not None and lg is not None:
                    sets.append((wg, lg))

            comment = str(d.get('Comment') or '').strip()
            if comment.lower() not in ('completed', ''):
                continue  # Skip walkovers, retirements

            matches.append({
                'winner': winner,
                'loser':  loser,
                'date':   date_str,
                'sets':   sets,
            })
        _XLSX_CACHE[path] = matches
        log.debug('Loaded %d matches from %s', len(matches), os.path.basename(path))
        return matches
    except Exception as e:
        log.warning('Failed to load %s: %s', path, e)
        return []


def _sim(a, b):
    a, b = a.lower().strip(), b.lower().strip()
    if a == b: return 1.0
    if a in b or b in a: return 0.85
    return SequenceMatcher(None, a, b).ratio()


def _last_name(full_name):
    """Extract last name from full name like 'Casper Ruud' → 'Ruud'."""
    parts = full_name.strip().split()
    return parts[-1] if parts else full_name


def _xlsx_last_name(xlsx_name):
    """Extract last name from XLSX format 'Ruud C.' → 'Ruud'."""
    # Handle double last names: "Bautista Agut R." → "Bautista Agut"
    m = re.match(r'^(.+?)\s+[A-Z]\.$', xlsx_name.strip())
    if m:
        return m.group(1).strip()
    return xlsx_name.strip()


def _name_sim(full_sib, xlsx_name):
    """Match sibila full name against XLSX 'Lastname F.' format."""
    sib_last = _last_name(full_sib).lower()
    xlsx_last = _xlsx_last_name(xlsx_name).lower()
    if sib_last == xlsx_last:
        return 1.0
    if sib_last in xlsx_last or xlsx_last in sib_last:
        return 0.9
    # Also try full fuzzy
    return _sim(full_sib.lower(), xlsx_name.lower())


def _find_match(all_rows, player1, player2, ts_str):
    """Find match involving player1 vs player2 within ±2 days of ts_str."""
    try:
        pick_dt = datetime.strptime(ts_str[:10], '%Y-%m-%d')
    except Exception:
        return None, None, None

    dates = [(pick_dt + timedelta(days=d)).strftime('%Y-%m-%d') for d in (-2, -1, 0, 1, 2)]

    best, best_score = None, 0.0
    best_p1_is_winner = None

    for row in all_rows:
        if row['date'] not in dates:
            continue
        w_sim1 = _name_sim(player1, row['winner'])
        w_sim2 = _name_sim(player2, row['winner'])
        l_sim1 = _name_sim(player1, row['loser'])
        l_sim2 = _name_sim(player2, row['loser'])

        # p1=winner, p2=loser
        score_a = min(w_sim1, l_sim2)
        # p1=loser, p2=winner
        score_b = min(l_sim1, w_sim2)

        if score_a >= score_b and score_a > best_score and score_a >= 0.5:
            best_score = score_a
            best = row
            best_p1_is_winner = True
        elif score_b > score_a and score_b > best_score and score_b >= 0.5:
            best_score = score_b
            best = row
            best_p1_is_winner = False

    if best is None:
        return None, None, None
    return best, best_p1_is_winner, best_score


def _parse_match(match_str):
    for sep in [' vs ', ' - ', ' v ']:
        if sep in match_str:
            p = match_str.split(sep, 1)
            return p[0].strip(), p[1].strip()
    return match_str.strip(), ''


def _parse_side(side_str, player1, player2):
    """
    Returns (kind, params) where kind is one of:
      'winner' → params = player_full_name (the predicted winner)
      'exact_sets' → params = N (int)
      'sets_ou' → params = (direction, line)
      'w_total' → params = (player_side, direction, line)
        player_side: 'home'|'away'
      'wins_set' → params = player_name_fragment
    """
    s = side_str.lower().strip()

    # Winner: Player Name
    m = re.match(r'winner:\s*(.+)', side_str, re.IGNORECASE)
    if m:
        return 'winner', m.group(1).strip()

    # Exact N sets
    m = re.match(r'exact\s+(\d+)\s+sets?', s)
    if m:
        return 'exact_sets', int(m.group(1))

    # Sets Under/Over X.5
    m = re.search(r'sets?\s+(under|over)\s+([\d.]+)', s)
    if m:
        return 'sets_ou', (m.group(1), float(m.group(2)))

    # W+Total home_and_over / away_and_under etc.
    m = re.search(r'w\+total\s+(home|away)_and_(over|under)\s+([\d.]+)', s)
    if m:
        return 'w_total', (m.group(1), m.group(2), float(m.group(3)))

    # X wins set (yes)
    m = re.match(r'(.+?)\s+wins?\s+set', side_str, re.IGNORECASE)
    if m:
        return 'wins_set', m.group(1).strip()

    return None, None


def _eval(row, p1_is_winner, kind, params, player1, player2):
    """Evaluate WIN/LOSS for a resolved match row."""
    sets = row['sets']
    n_sets = len(sets)
    total_games = sum(w + l for w, l in sets)

    if kind == 'winner':
        predicted = params
        # Match predicted name against p1/p2
        sim1 = _sim(predicted.lower(), player1.lower())
        sim2 = _sim(predicted.lower(), player2.lower())
        if sim1 < 0.4 and sim2 < 0.4:
            # Try last name match
            pred_last = _last_name(predicted).lower()
            sim1 = _sim(pred_last, _last_name(player1).lower())
            sim2 = _sim(pred_last, _last_name(player2).lower())
        p1_picked = sim1 >= sim2
        won = (p1_picked and p1_is_winner) or (not p1_picked and not p1_is_winner)
        return ('WIN' if won else 'LOSS'), f'winner={row["winner"]} sets={n_sets}'

    if kind == 'exact_sets':
        target = params
        won = (n_sets == target)
        return ('WIN' if won else 'LOSS'), f'n_sets={n_sets} target={target}'

    if kind == 'sets_ou':
        direction, line = params
        if direction == 'under':
            won = n_sets < line
        else:
            won = n_sets > line
        if n_sets == line:
            return 'VOID', f'n_sets={n_sets}==line={line}'
        return ('WIN' if won else 'LOSS'), f'n_sets={n_sets} {direction} {line}'

    if kind == 'w_total':
        player_side, direction, line = params
        # p1 = home, p2 = away
        if player_side == 'home':
            player_wins = p1_is_winner
        else:
            player_wins = not p1_is_winner

        if not player_wins:
            return 'LOSS', f'wrong_winner total_games={total_games}'

        if direction == 'over':
            result_total = 'WIN' if total_games > line else 'LOSS'
        else:
            result_total = 'WIN' if total_games < line else 'LOSS'

        if total_games == line:
            return 'VOID', f'total_games={total_games}==line'
        return result_total, f'player_wins=True total={total_games} {direction} {line}'

    if kind == 'wins_set':
        frag = params.lower()
        # Determine which player the fragment refers to
        sim1 = _sim(frag, player1.lower()) + _sim(frag, _last_name(player1).lower())
        sim2 = _sim(frag, player2.lower()) + _sim(frag, _last_name(player2).lower())
        refers_to_p1 = sim1 >= sim2
        refers_to_winner = (refers_to_p1 and p1_is_winner) or (not refers_to_p1 and not p1_is_winner)

        if refers_to_winner:
            won = True  # Winner always wins at least one set
        else:
            # Loser wins a set if they won any set
            loser_won_set = any(l > w for w, l in sets)
            won = loser_won_set
        return ('WIN' if won else 'LOSS'), f'ref_p1={refers_to_p1} won_set={won}'

    return None, None


def resolve_all_pending(dry_run=False):
    all_rows = _load_xlsx(ATP_XLSX) + _load_xlsx(WTA_XLSX)
    log.info('Loaded %d tennis match records from XLSX', len(all_rows))

    sib_conn = sqlite3.connect(SIBILA_DB)
    cols = [r[1] for r in sib_conn.execute('PRAGMA table_info(sibila_picks)').fetchall()]

    rows = sib_conn.execute(
        "SELECT * FROM sibila_picks WHERE sport='tennis' "
        "AND (result IS NULL OR result='') ORDER BY ts"
    ).fetchall()

    log.info('Pending tennis picks: %d', len(rows))

    def col(row, name):
        idx = cols.index(name) if name in cols else None
        return row[idx] if idx is not None else None

    resolved = not_found = skipped = 0

    for row in rows:
        pick_id = col(row, 'id')
        match   = str(col(row, 'match') or '')
        side    = str(col(row, 'side')  or '')
        ts      = str(col(row, 'ts')    or '')
        odds       = float(col(row, 'odds') or 1.5)
        stake      = float(col(row, 'shadow_stake') or col(row, 'stake') or 10.0)
        event_id   = str(col(row, 'event_id') or '')
        market_url = str(col(row, 'market_url') or '')

        player1, player2 = _parse_match(match)
        if not player2:
            skipped += 1
            continue

        kind, params = _parse_side(side, player1, player2)
        if kind is None:
            log.warning('  Cannot parse: %s | %s', match, side)
            skipped += 1
            continue

        game_row, p1_is_winner, score = _find_match(all_rows, player1, player2, ts)
        if game_row is None:
            log.debug('  Not found: %s (%s)', match, ts[:10])
            not_found += 1
            continue

        result, detail = _eval(game_row, p1_is_winner, kind, params, player1, player2)
        if result is None:
            skipped += 1
            continue

        pnl = (odds - 1) * stake if result == 'WIN' else (-stake if result == 'LOSS' else 0.0)
        log.info('  %s | %s | %s | %s | $%+.2f', result, match[:30], side[:25], detail, pnl)

        if not dry_run:
            now = datetime.utcnow().isoformat()
            closing_odds = None
            if event_id and market_url:
                try:
                    import sys as _sys, os as _os
                    _sys.path.insert(0, _os.path.dirname(_os.path.abspath(__file__)))
                    from oraculo_odds_monitor import get_closing_odds as _gco
                    closing_odds = _gco(event_id, market_url)
                except Exception:
                    pass
            clv = None
            if closing_odds and closing_odds > 1 and odds > 1:
                clv = round(closing_odds / odds - 1.0, 4)
            sib_conn.execute(
                "UPDATE sibila_picks "
                "SET result=?, pnl=?, resolved_ts=?, "
                "closing_odds=COALESCE(closing_odds,?), clv=COALESCE(clv,?) "
                "WHERE id=?",
                (result, pnl, now, closing_odds, clv, pick_id)
            )
            sib_conn.commit()
            resolved += 1

    # Summary
    settled = sib_conn.execute(
        "SELECT result, COUNT(*), SUM(pnl) FROM sibila_picks "
        "WHERE sport='tennis' AND result IN ('WIN','LOSS','VOID') GROUP BY result"
    ).fetchall()
    sib_conn.close()

    print('\n=== Tennis Shadow Summary ===')
    tn = tw = 0
    tpnl = 0.0
    for res, n, pnl in settled:
        pnl = pnl or 0
        print(f'  {res:4}: n={n}  PnL=${pnl:+.2f}')
        tn += n
        if res == 'WIN': tw += n
        tpnl += pnl
    if tn:
        print(f'  TOTAL: {tw}/{tn}  WR={tw/tn*100:.1f}%  PnL=${tpnl:+.2f}')
    print(f'\nThis run: resolved={resolved}  skipped={skipped}  not_found={not_found}')
    return resolved, skipped, not_found


if __name__ == '__main__':
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument('--dry-run', action='store_true')
    args = p.parse_args()
    resolve_all_pending(dry_run=args.dry_run)
